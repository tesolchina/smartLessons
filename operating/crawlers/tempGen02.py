import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Create a new workbook
wb = openpyxl.Workbook()

# Define tab names (replace invalid '/' with '-')
tabs = [
    "Overview",
    "Front Page Template",
    "Unified Template (Events-Courses-Services)",  # Replace '/' with '-'
    "Staff Template"
]

# Create tabs
for tab in tabs:
    if tab == "Overview":
        ws = wb.active
        ws.title = tab
    else:
        wb.create_sheet(title=tab)

# Populate "Overview" tab
overview = wb["Overview"]
overview.append(["Tab Name", "Description", "Link to Tab"])
for idx, tab in enumerate(tabs[1:], start=1):
    overview.append([tab, f"Description for {tab}", f"=HYPERLINK(\"#{tab}!A1\", \"Go to {tab}\")"])
overview.column_dimensions[get_column_letter(1)].width = 30
overview.column_dimensions[get_column_letter(2)].width = 50
overview.column_dimensions[get_column_letter(3)].width = 20
header_font = Font(bold=True)
for cell in overview[1]:
    cell.font = header_font

# Populate other tabs with updated templates
templates = {
    "Front Page Template": ["Title", "Introduction", "Key Content", "Images/Media"],
    "Unified Template (Events-Courses-Services)": [  # Updated tab name
        "Title", "Description", "Date/Time", "Location/Link", "Organizer Name/Contact",
        "Banner Image(s)", "Registration Link/Form", "Registration Deadline", 
        "Available Slots/Capacity", "Visitor Stories/Testimonials", "Media (Images/Videos)", 
        "Event Highlights", "Social Media Handles"
    ],
    "Staff Template": ["Name", "Position", "Contact Info", "Bio", "Photo"]
}

for tab, fields in templates.items():
    ws = wb[tab]
    ws.append(fields)
    for col_num, _ in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    for cell in ws[1]:
        cell.font = header_font

# Save the workbook with the new file name
wb.save("LC_Website_Migration_Templates02.xlsx")
print("Spreadsheet created successfully with the new name!")