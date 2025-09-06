import os
from docx import Document
from openpyxl import Workbook

# Define the folder structure
structure = {
    "About Us": ["Mission", "Staff", "Job Vacancies", "Contact Us"],
    "Courses": ["English", "Chinese", "Putonghua", "Foreign Languages"],
    "Self-regulated Language Learning": ["Examination", "Tutorial Services", "Activities", "Professional Development"],
    "Staff Corner": ["Leadership", "Research", "Staff Development", "Social"],
}

# Base directory (use a shorter path to avoid long path issues)
base_dir = r"C:\LC_Planning"
os.makedirs(base_dir, exist_ok=True)

# Helper function to sanitize folder names
def sanitize_name(name):
    return name.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace('"', "_").replace("<", "_").replace(">", "_").replace("|", "_")

# Create Excel workbook
wb = Workbook()

# Iterate over main folders
for main_folder, sub_folders in structure.items():
    # Sanitize main folder name
    sanitized_main_folder = sanitize_name(main_folder)
    
    # Create main folder
    main_folder_path = os.path.join(base_dir, sanitized_main_folder)
    print(f"Creating folder: {main_folder_path}")  # Debugging line
    os.makedirs(main_folder_path, exist_ok=True)
    
    # Create a new sheet in the Excel workbook
    sheet = wb.create_sheet(title=sanitized_main_folder[:31])  # Excel sheet titles have a 31-character limit
    sheet.append(["Sub-folder Name", "Word Document Link", "HTML Template Link"])
    
    # Iterate over sub-folders
    for sub_folder in sub_folders:
        # Sanitize sub-folder name
        sanitized_sub_folder = sanitize_name(sub_folder)
        
        # Create sub-folder
        sub_folder_path = os.path.join(main_folder_path, sanitized_sub_folder)
        print(f"Creating sub-folder: {sub_folder_path}")  # Debugging line
        os.makedirs(sub_folder_path, exist_ok=True)
        
        # Create Word document
        word_doc_name = f"{sanitized_sub_folder}_planning.docx"
        word_doc_path = os.path.join(sub_folder_path, word_doc_name)
        doc = Document()
        doc.add_heading(f"{sub_folder} Planning Document", level=1)
        doc.save(word_doc_path)
        
        # Create HTML template
        html_template_name = f"{sanitized_sub_folder}_template.html"
        html_template_path = os.path.join(sub_folder_path, html_template_name)
        with open(html_template_path, "w") as html_file:
            html_file.write(f"<!DOCTYPE html>\n<html>\n<head>\n<title>{sub_folder} Template</title>\n</head>\n<body>\n<h1>{sub_folder}</h1>\n</body>\n</html>")
        
        # Add hyperlinks to the Excel sheet
        word_link = f'=HYPERLINK("{word_doc_path}", "Word Doc")'
        html_link = f'=HYPERLINK("{html_template_path}", "HTML Template")'
        sheet.append([sub_folder, word_link, html_link])

# Remove the default sheet
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Save the Excel workbook
excel_path = os.path.join(base_dir, "Language_Centre_Structure.xlsx")
wb.save(excel_path)

print(f"Folder structure, files, and Excel sheet created in '{base_dir}'")
