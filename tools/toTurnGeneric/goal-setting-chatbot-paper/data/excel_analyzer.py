#!/usr/bin/env python3
"""
Excel Data Reader with xlrd fallback
Handles Excel files with older openpyxl compatibility
Created: September 6, 2025
"""

import pandas as pd
import matplotlib.pyplot as plt
import os
import warnings
warnings.filterwarnings('ignore')

def load_excel_with_fallback(file_path):
    """Try multiple methods to load Excel file"""
    print(f"📊 Loading Excel file: {file_path}")
    
    # Method 1: Try openpyxl engine
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        print("✅ Loaded with openpyxl engine")
        return df
    except Exception as e:
        print(f"⚠️ openpyxl failed: {e}")
    
    # Method 2: Try xlrd engine  
    try:
        df = pd.read_excel(file_path, engine='xlrd')
        print("✅ Loaded with xlrd engine")
        return df
    except Exception as e:
        print(f"⚠️ xlrd failed: {e}")
    
    # Method 3: Try without specifying engine
    try:
        df = pd.read_excel(file_path)
        print("✅ Loaded with default engine")
        return df
    except Exception as e:
        print(f"⚠️ Default engine failed: {e}")
    
    # Method 4: Try converting to CSV first
    try:
        print("🔄 Trying alternative approach...")
        # This would require manual conversion, let's skip for now
        return None
    except:
        pass
    
    print("❌ All methods failed")
    return None

def examine_excel_structure(file_path):
    """Examine Excel file structure without pandas"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        print(f"📋 Workbook contains {len(wb.sheetnames)} sheet(s):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
            
            ws = wb[sheet_name]
            print(f"      Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Get header row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else "")
            print(f"      Headers: {headers[:8]}{'...' if len(headers) > 8 else ''}")
            
            # Sample first few data rows
            print("      Sample data:")
            for row_num in range(2, min(5, ws.max_row + 1)):
                row_data = []
                for cell in ws[row_num]:
                    row_data.append(str(cell.value) if cell.value else "")
                print(f"        Row {row_num}: {row_data[:5]}{'...' if len(row_data) > 5 else ''}")
            
            if i == 1:  # Only show details for first sheet
                return headers, ws.max_row, ws.max_column
            
    except Exception as e:
        print(f"❌ Error examining Excel structure: {e}")
        return None, 0, 0

def manual_excel_analysis(file_path):
    """Manually read and analyze Excel data"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active  # Get the active sheet
        
        print(f"📊 Manually processing {ws.max_row} rows...")
        
        # Read header row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")
        
        print(f"Headers: {headers}")
        
        # Find relevant columns
        student_col_idx = None
        content_col_idx = None
        time_col_idx = None
        
        for i, header in enumerate(headers):
            header_lower = str(header).lower()
            if any(word in header_lower for word in ['student', 'user', 'name', 'id']):
                student_col_idx = i
            elif any(word in header_lower for word in ['content', 'message', 'chat', 'text']):
                content_col_idx = i
            elif any(word in header_lower for word in ['time', 'date', 'created']):
                time_col_idx = i
        
        print(f"📍 Found columns - Student: {student_col_idx}, Content: {content_col_idx}, Time: {time_col_idx}")
        
        if student_col_idx is None:
            print("❌ Could not find student column")
            return None
        
        # Read data
        student_stats = {}
        
        for row_num in range(2, ws.max_row + 1):
            row = list(ws[row_num])
            
            # Get student ID
            student_id = str(row[student_col_idx].value) if row[student_col_idx].value else ""
            if not student_id or student_id == "None":
                continue
                
            # Initialize student stats
            if student_id not in student_stats:
                student_stats[student_id] = {
                    'chat_rounds': 0,
                    'total_words': 0,
                    'dates': set()
                }
            
            # Count chat round
            student_stats[student_id]['chat_rounds'] += 1
            
            # Count words if content column exists
            if content_col_idx is not None:
                content = str(row[content_col_idx].value) if row[content_col_idx].value else ""
                words = len(content.split()) if content != "None" else 0
                student_stats[student_id]['total_words'] += words
            
            # Track dates if time column exists
            if time_col_idx is not None:
                date_val = row[time_col_idx].value
                if date_val:
                    if hasattr(date_val, 'date'):
                        student_stats[student_id]['dates'].add(date_val.date())
                    else:
                        student_stats[student_id]['dates'].add(str(date_val)[:10])
        
        # Convert to results format
        results = []
        for student_id, stats in student_stats.items():
            results.append({
                'Student': student_id,
                'Login_Count': len(stats['dates']) if stats['dates'] else 1,
                'Chat_Rounds': stats['chat_rounds'],
                'Total_Words': stats['total_words']
            })
        
        # Convert to DataFrame
        results_df = pd.DataFrame(results)
        
        print(f"✅ Processed {len(results_df)} students")
        return results_df
        
    except Exception as e:
        print(f"❌ Manual analysis failed: {e}")
        return None

def create_simple_charts(results_df, output_dir):
    """Create charts with matplotlib only"""
    print("📈 Creating charts...")
    
    # Create main analysis chart
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Chatbot Usage Analysis - Goal Setting Study', fontsize=16, fontweight='bold')
    
    students = results_df['Student'].tolist()
    x_pos = range(len(students))
    
    # 1. Login counts
    ax1.bar(x_pos, results_df['Login_Count'], color='skyblue', alpha=0.7)
    ax1.set_title('Login Count per Student')
    ax1.set_xlabel('Students')
    ax1.set_ylabel('Number of Logins')
    ax1.grid(True, alpha=0.3)
    
    # 2. Chat rounds
    ax2.bar(x_pos, results_df['Chat_Rounds'], color='lightgreen', alpha=0.7)
    ax2.set_title('Chat Rounds per Student')
    ax2.set_xlabel('Students')
    ax2.set_ylabel('Number of Chat Rounds')
    ax2.grid(True, alpha=0.3)
    
    # 3. Word counts
    ax3.bar(x_pos, results_df['Total_Words'], color='salmon', alpha=0.7)
    ax3.set_title('Total Words per Student')
    ax3.set_xlabel('Students')
    ax3.set_ylabel('Total Word Count')
    ax3.grid(True, alpha=0.3)
    
    # 4. Scatter plot
    ax4.scatter(results_df['Chat_Rounds'], results_df['Total_Words'], color='purple', alpha=0.6)
    ax4.set_title('Chat Rounds vs Total Words')
    ax4.set_xlabel('Chat Rounds')
    ax4.set_ylabel('Total Words')
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    chart_path = os.path.join(output_dir, 'chatbot_analysis_results.png')
    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
    print(f"✅ Chart saved: {chart_path}")
    
    plt.close()
    
    # Create summary statistics
    plt.figure(figsize=(10, 6))
    
    categories = ['Avg Logins', 'Avg Chat Rounds', 'Avg Words']
    values = [
        results_df['Login_Count'].mean(),
        results_df['Chat_Rounds'].mean(),
        results_df['Total_Words'].mean()
    ]
    
    bars = plt.bar(categories, values, color=['skyblue', 'lightgreen', 'salmon'], alpha=0.7)
    plt.title('Average Usage Statistics', fontsize=14, fontweight='bold')
    plt.ylabel('Average Count')
    plt.grid(True, alpha=0.3)
    
    # Add values on bars
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height, f'{height:.1f}',
                ha='center', va='bottom')
    
    plt.tight_layout()
    
    summary_path = os.path.join(output_dir, 'usage_summary.png')
    plt.savefig(summary_path, dpi=300, bbox_inches='tight')
    print(f"✅ Summary saved: {summary_path}")
    
    plt.close()

def main():
    """Main execution"""
    print("🤖 Chatbot Data Analyzer (Excel Compatible)")
    print("="*50)
    
    base_dir = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper"
    excel_file = os.path.join(base_dir, "data", "final_report_fulfill.xlsx")
    csv_output = os.path.join(base_dir, "data", "usage_analysis_results.csv")
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    # First, examine the Excel structure
    print("\n🔍 Examining Excel structure...")
    headers, max_row, max_col = examine_excel_structure(excel_file)
    
    if max_row == 0:
        print("❌ Could not examine Excel file")
        return
    
    print(f"\n📊 File contains {max_row} rows and {max_col} columns")
    
    # Try pandas first, then fallback to manual processing
    results_df = load_excel_with_fallback(excel_file)
    
    if results_df is None:
        print("\n🔧 Falling back to manual Excel processing...")
        results_df = manual_excel_analysis(excel_file)
    else:
        # If pandas worked, do the regular analysis
        print("\n📈 Analyzing with pandas...")
        # Quick analysis since pandas worked
        print(f"Loaded {len(results_df)} rows")
        print("Columns:", list(results_df.columns))
    
    if results_df is None:
        print("❌ Failed to process Excel file")
        return
    
    # Display results
    print(f"\n📊 ANALYSIS RESULTS")
    print(f"Students analyzed: {len(results_df)}")
    if len(results_df) > 0:
        print(f"Average logins: {results_df['Login_Count'].mean():.2f}")
        print(f"Average chat rounds: {results_df['Chat_Rounds'].mean():.2f}")
        print(f"Average words: {results_df['Total_Words'].mean():.2f}")
        
        # Top students
        print(f"\n🏆 Top 5 Students by Chat Activity:")
        top_5 = results_df.nlargest(5, 'Chat_Rounds')
        for _, row in top_5.iterrows():
            print(f"   {row['Student']}: {row['Login_Count']} logins, "
                  f"{row['Chat_Rounds']} rounds, {row['Total_Words']} words")
    
    # Save results
    try:
        results_df.to_csv(csv_output, index=False)
        print(f"✅ CSV saved: {csv_output}")
    except Exception as e:
        print(f"❌ Error saving CSV: {e}")
    
    # Create charts
    try:
        create_simple_charts(results_df, os.path.join(base_dir, "data"))
        print(f"✅ Charts created in: {base_dir}/data/")
    except Exception as e:
        print(f"❌ Error creating charts: {e}")
    
    print(f"\n🎉 Analysis complete!")

if __name__ == "__main__":
    main()
