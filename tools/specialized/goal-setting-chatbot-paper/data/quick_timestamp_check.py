#!/usr/bin/env python3
"""
Quick Timestamp Inspector
Examines a few sample timestamp entries to understand the data structure
"""

import openpyxl
import json
from datetime import datetime

def quick_timestamp_check():
    """Quick examination of timestamp patterns"""
    
    print("🔍 QUICK TIMESTAMP INSPECTION")
    print("="*50)
    
    excel_file = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper/data/final_report_fulfill.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    ws = wb.active
    
    print(f"Examining first 5 students' timestamp patterns...")
    
    for row_num in range(2, min(7, ws.max_row + 1)):  # Check first 5 students only
        row = list(ws[row_num])
        
        student_id = str(row[0].value) if row[0].value else f"Student_{row_num-1}"
        dates_raw = str(row[2].value) if row[2].value else "[]"
        chat_rounds = row[5].value if row[5].value else 0
        
        print(f"\n👤 Student {student_id}")
        print(f"   Chat rounds: {chat_rounds}")
        print(f"   Raw timestamp: {dates_raw[:150]}...")
        
        # Try to parse timestamps
        try:
            dates_clean = dates_raw.replace("'", '"')
            if dates_clean.startswith('[') and dates_clean.endswith(']'):
                dates_list = json.loads(dates_clean)
                print(f"   📅 Number of timestamps: {len(dates_list)}")
                
                if len(dates_list) > 1:
                    print(f"   🕐 Timestamp analysis:")
                    timestamps = []
                    for i, date_str in enumerate(dates_list[:10]):  # First 10 only
                        try:
                            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            timestamps.append(dt)
                            print(f"      {i+1}: {dt.strftime('%Y-%m-%d %H:%M:%S')}")
                        except Exception as e:
                            print(f"      {i+1}: Error parsing - {e}")
                    
                    # Check time gaps
                    if len(timestamps) > 1:
                        timestamps.sort()
                        gaps = []
                        for i in range(1, len(timestamps)):
                            gap = (timestamps[i] - timestamps[i-1]).total_seconds() / 60
                            gaps.append(gap)
                            print(f"      Gap {i}: {gap:.1f} minutes")
                        
                        print(f"   📊 Summary: {len(timestamps)} timestamps, gaps: {min(gaps):.1f}-{max(gaps):.1f} min")
                
                elif len(dates_list) == 1:
                    print(f"   📅 Single timestamp: {dates_list[0]}")
                    
        except Exception as e:
            print(f"   ❌ Error parsing: {e}")

if __name__ == "__main__":
    quick_timestamp_check()
