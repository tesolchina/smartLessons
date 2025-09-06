#!/usr/bin/env python3
"""
Detailed Timestamp Analysis for Chatbot Study
Analyzes chat timestamps to identify real login sessions and interaction patterns
Created: September 6, 2025
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import json
import re
import os
from datetime import datetime, timedelta
import matplotlib.dates as mdates

def analyze_timestamps_detailed():
    """Detailed analysis of chat timestamps to identify login sessions"""
    
    print("🕐 DETAILED TIMESTAMP ANALYSIS")
    print("="*60)
    
    base_dir = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper"
    excel_file = os.path.join(base_dir, "data", "final_report_fulfill.xlsx")
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        print(f"📊 Analyzing timestamps for {ws.max_row-1} students...")
        
        detailed_results = []
        
        for row_num in range(2, ws.max_row + 1):
            row = list(ws[row_num])
            
            student_id = str(row[0].value) if row[0].value else f"Student_{row_num-1}"
            email = str(row[1].value) if row[1].value else ""
            dates_raw = str(row[2].value) if row[2].value else "[]"
            avg_user_words = row[3].value if row[3].value else 0
            avg_ai_words = row[4].value if row[4].value else 0
            chat_rounds = row[5].value if row[5].value else 0
            
            print(f"\n👤 Student {student_id[:15]}...")
            print(f"   Email: {email}")
            print(f"   Chat rounds: {chat_rounds}")
            print(f"   Raw dates: {dates_raw[:100]}...")
            
            # Parse timestamps more carefully
            timestamps = []
            
            try:
                # Clean up the timestamp string
                dates_clean = dates_raw.replace("'", '"')
                
                if dates_clean.startswith('[') and dates_clean.endswith(']'):
                    # Parse as JSON
                    dates_list = json.loads(dates_clean)
                    print(f"   📅 Found {len(dates_list)} timestamp entries")
                    
                    for i, date_str in enumerate(dates_list):
                        try:
                            # Parse ISO format timestamp
                            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                            timestamps.append(dt)
                            if i < 5:  # Show first 5 timestamps
                                print(f"      {i+1}. {dt.strftime('%Y-%m-%d %H:%M:%S')}")
                        except Exception as e:
                            print(f"      ❌ Error parsing timestamp {i+1}: {e}")
                    
                    if len(dates_list) > 5:
                        print(f"      ... and {len(dates_list)-5} more timestamps")
                        
                else:
                    print(f"   ⚠️ Unexpected timestamp format")
                    
            except Exception as e:
                print(f"   ❌ Error parsing timestamps: {e}")
                
            # Analyze timestamp patterns if we have multiple timestamps
            login_sessions = 1  # Default
            time_gaps = []
            session_details = []
            
            if len(timestamps) > 1:
                # Sort timestamps
                timestamps.sort()
                
                print(f"   🕐 Timeline analysis:")
                print(f"      First: {timestamps[0].strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"      Last:  {timestamps[-1].strftime('%Y-%m-%d %H:%M:%S')}")
                
                # Calculate time gaps between consecutive interactions
                for i in range(1, len(timestamps)):
                    gap = timestamps[i] - timestamps[i-1]
                    time_gaps.append(gap.total_seconds() / 60)  # Convert to minutes
                
                print(f"      Time gaps (minutes): {[f'{g:.1f}' for g in time_gaps[:10]]}")
                
                # Identify sessions (gaps > 30 minutes indicate new session)
                SESSION_THRESHOLD = 30  # minutes
                
                current_session_start = timestamps[0]
                current_session_chats = 1
                sessions = []
                
                for i, gap in enumerate(time_gaps):
                    if gap > SESSION_THRESHOLD:
                        # End current session
                        sessions.append({
                            'start': current_session_start,
                            'end': timestamps[i],
                            'chats': current_session_chats,
                            'duration_minutes': (timestamps[i] - current_session_start).total_seconds() / 60
                        })
                        
                        # Start new session
                        current_session_start = timestamps[i+1]
                        current_session_chats = 1
                    else:
                        current_session_chats += 1
                
                # Add final session
                sessions.append({
                    'start': current_session_start,
                    'end': timestamps[-1],
                    'chats': current_session_chats,
                    'duration_minutes': (timestamps[-1] - current_session_start).total_seconds() / 60
                })
                
                login_sessions = len(sessions)
                
                print(f"   📊 Session Analysis (gaps > {SESSION_THRESHOLD} min = new session):")
                print(f"      Number of sessions: {login_sessions}")
                
                for i, session in enumerate(sessions, 1):
                    print(f"      Session {i}: {session['chats']} chats, "
                          f"{session['duration_minutes']:.1f} min duration")
                    print(f"                {session['start'].strftime('%Y-%m-%d %H:%M')} → "
                          f"{session['end'].strftime('%H:%M')}")
                
                session_details = sessions
            
            # Calculate total user words
            try:
                total_user_words = int(float(chat_rounds) * float(avg_user_words)) if chat_rounds and avg_user_words else 0
            except:
                total_user_words = 0
            
            detailed_results.append({
                'Student_ID': student_id,
                'Email': email,
                'Original_Login_Count': 1,  # Original assumption
                'Calculated_Login_Sessions': login_sessions,
                'Chat_Rounds': int(float(chat_rounds)) if chat_rounds else 0,
                'Avg_User_Words': round(float(avg_user_words), 2) if avg_user_words else 0,
                'Avg_AI_Words': round(float(avg_ai_words), 2) if avg_ai_words else 0,
                'Total_User_Words': total_user_words,
                'Total_Timestamps': len(timestamps),
                'First_Chat': timestamps[0].strftime('%Y-%m-%d %H:%M:%S') if timestamps else '',
                'Last_Chat': timestamps[-1].strftime('%Y-%m-%d %H:%M:%S') if timestamps else '',
                'Study_Duration_Hours': (timestamps[-1] - timestamps[0]).total_seconds() / 3600 if len(timestamps) > 1 else 0,
                'Max_Gap_Minutes': max(time_gaps) if time_gaps else 0,
                'Min_Gap_Minutes': min(time_gaps) if time_gaps else 0,
                'Avg_Gap_Minutes': np.mean(time_gaps) if time_gaps else 0,
                'Session_Details': str(session_details)[:200] + "..." if len(str(session_details)) > 200 else str(session_details)
            })
        
        # Convert to DataFrame
        df = pd.DataFrame(detailed_results)
        
        # Filter out students with no chat activity
        df = df[df['Chat_Rounds'] > 0]
        
        print(f"\n📈 REVISED ANALYSIS SUMMARY")
        print(f"={'='*50}")
        print(f"Total active students: {len(df)}")
        print(f"Original login assumption (all=1): {df['Original_Login_Count'].sum()}")
        print(f"Calculated login sessions: {df['Calculated_Login_Sessions'].sum()}")
        print(f"Total chat rounds: {df['Chat_Rounds'].sum()}")
        print(f"Students with multiple sessions: {len(df[df['Calculated_Login_Sessions'] > 1])}")
        print(f"Max sessions by one student: {df['Calculated_Login_Sessions'].max()}")
        print(f"Average sessions per student: {df['Calculated_Login_Sessions'].mean():.2f}")
        
        # Show students with multiple sessions
        multi_session = df[df['Calculated_Login_Sessions'] > 1]
        if len(multi_session) > 0:
            print(f"\n🔍 STUDENTS WITH MULTIPLE SESSIONS:")
            for _, row in multi_session.iterrows():
                print(f"   {row['Student_ID'][:15]}: {row['Calculated_Login_Sessions']} sessions, "
                      f"{row['Chat_Rounds']} chats, {row['Study_Duration_Hours']:.1f}h span")
        else:
            print(f"\n⚠️  NO STUDENTS WITH MULTIPLE SESSIONS DETECTED")
            print(f"   This suggests either:")
            print(f"   1. All interactions were in single continuous sessions")
            print(f"   2. The timestamp data doesn't contain multiple session info")
            print(f"   3. The session threshold (30 min) needs adjustment")
        
        return df
        
    except Exception as e:
        print(f"❌ Error in detailed analysis: {e}")
        return None

def create_timeline_visualization(df, output_dir):
    """Create timeline visualizations showing interaction patterns"""
    
    print(f"\n📊 Creating timeline visualizations...")
    
    # 1. Sessions comparison chart
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Detailed Chatbot Interaction Analysis', fontsize=16, fontweight='bold')
    
    # Original vs Calculated Sessions
    x = range(len(df))
    ax1.bar([i-0.2 for i in x], df['Original_Login_Count'], 0.4, 
            label='Original (assumed)', color='lightcoral', alpha=0.7)
    ax1.bar([i+0.2 for i in x], df['Calculated_Login_Sessions'], 0.4,
            label='Calculated', color='skyblue', alpha=0.7)
    ax1.set_title('Login Sessions: Original vs Calculated')
    ax1.set_xlabel('Student Index')
    ax1.set_ylabel('Number of Sessions')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Study duration vs chat rounds
    ax2.scatter(df['Study_Duration_Hours'], df['Chat_Rounds'], 
                color='green', alpha=0.6, s=60)
    ax2.set_title('Study Duration vs Chat Rounds')
    ax2.set_xlabel('Duration (Hours)')
    ax2.set_ylabel('Chat Rounds')
    ax2.grid(True, alpha=0.3)
    
    # Add correlation
    if len(df) > 1:
        corr = df['Study_Duration_Hours'].corr(df['Chat_Rounds'])
        ax2.text(0.05, 0.95, f'Correlation: {corr:.3f}', 
                transform=ax2.transAxes, fontweight='bold',
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # Gap analysis
    ax3.bar(range(len(df)), df['Max_Gap_Minutes'], color='orange', alpha=0.7)
    ax3.set_title('Maximum Gap Between Chats (Minutes)')
    ax3.set_xlabel('Student Index')
    ax3.set_ylabel('Max Gap (Minutes)')
    ax3.grid(True, alpha=0.3)
    
    # Sessions distribution
    session_counts = df['Calculated_Login_Sessions'].value_counts().sort_index()
    ax4.bar(session_counts.index, session_counts.values, color='purple', alpha=0.7)
    ax4.set_title('Distribution of Login Sessions')
    ax4.set_xlabel('Number of Sessions')
    ax4.set_ylabel('Number of Students')
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    timeline_chart = os.path.join(output_dir, 'detailed_timeline_analysis.png')
    plt.savefig(timeline_chart, dpi=300, bbox_inches='tight')
    print(f"✅ Timeline analysis saved: {timeline_chart}")
    plt.close()
    
    # 2. Create interaction pattern heatmap if we have time data
    fig, ax = plt.subplots(1, 1, figsize=(14, 8))
    
    # Summary statistics
    stats_data = {
        'Total Students': len(df),
        'Total Chat Rounds': df['Chat_Rounds'].sum(),
        'Total Calculated Sessions': df['Calculated_Login_Sessions'].sum(),
        'Avg Sessions/Student': df['Calculated_Login_Sessions'].mean(),
        'Avg Duration (Hours)': df['Study_Duration_Hours'].mean(),
        'Avg Max Gap (Minutes)': df['Max_Gap_Minutes'].mean(),
        'Students w/ Multiple Sessions': len(df[df['Calculated_Login_Sessions'] > 1])
    }
    
    # Horizontal bar chart
    y_pos = np.arange(len(stats_data))
    values = list(stats_data.values())
    labels = list(stats_data.keys())
    
    bars = ax.barh(y_pos, values, color=['skyblue', 'lightgreen', 'salmon', 'gold', 
                                         'lightcoral', 'plum', 'lightgray'], alpha=0.7)
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels)
    ax.set_xlabel('Count / Average')
    ax.set_title('Detailed Usage Statistics Summary', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    
    # Add value labels
    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax.text(width + max(values)*0.01, bar.get_y() + bar.get_height()/2, 
                f'{width:.2f}', ha='left', va='center', fontweight='bold')
    
    plt.tight_layout()
    
    summary_chart = os.path.join(output_dir, 'detailed_summary_statistics.png')
    plt.savefig(summary_chart, dpi=300, bbox_inches='tight')
    print(f"✅ Summary statistics saved: {summary_chart}")
    plt.close()

def save_detailed_results(df, output_dir):
    """Save detailed analysis results"""
    
    # Main detailed results
    detailed_csv = os.path.join(output_dir, 'detailed_timestamp_analysis.csv')
    df.to_csv(detailed_csv, index=False)
    print(f"✅ Detailed results saved: {detailed_csv}")
    
    # Comparison summary
    comparison_data = {
        'Analysis_Method': ['Original Assumption', 'Calculated from Timestamps'],
        'Total_Login_Sessions': [df['Original_Login_Count'].sum(), df['Calculated_Login_Sessions'].sum()],
        'Average_Sessions_Per_Student': [df['Original_Login_Count'].mean(), df['Calculated_Login_Sessions'].mean()],
        'Students_With_Multiple_Sessions': [0, len(df[df['Calculated_Login_Sessions'] > 1])],
        'Max_Sessions_One_Student': [1, df['Calculated_Login_Sessions'].max()]
    }
    
    comparison_df = pd.DataFrame(comparison_data)
    comparison_csv = os.path.join(output_dir, 'login_sessions_comparison.csv')
    comparison_df.to_csv(comparison_csv, index=False)
    print(f"✅ Comparison analysis saved: {comparison_csv}")

def main():
    """Main execution"""
    
    # Run detailed timestamp analysis
    df = analyze_timestamps_detailed()
    
    if df is None or len(df) == 0:
        print("❌ No data to analyze")
        return
    
    # Save results
    output_dir = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper/data"
    save_detailed_results(df, output_dir)
    
    # Create visualizations
    create_timeline_visualization(df, output_dir)
    
    print(f"\n🔍 DETAILED ANALYSIS CONCLUSIONS:")
    print(f"="*50)
    
    multi_session_count = len(df[df['Calculated_Login_Sessions'] > 1])
    if multi_session_count == 0:
        print(f"📝 Finding: All students appear to have used the chatbot in single sessions")
        print(f"   • This could indicate:")
        print(f"     - Students completed their interactions in one sitting")
        print(f"     - The study design encouraged single-session use")
        print(f"     - Technical limitations prevented session resumption")
        print(f"     - Data collection captured only active chat periods")
    else:
        print(f"📝 Finding: {multi_session_count} students had multiple login sessions")
        print(f"   • Average sessions per student: {df['Calculated_Login_Sessions'].mean():.2f}")
        print(f"   • Maximum sessions by one student: {df['Calculated_Login_Sessions'].max()}")
    
    avg_duration = df['Study_Duration_Hours'].mean()
    print(f"\n⏱️ Temporal Patterns:")
    print(f"   • Average study span: {avg_duration:.2f} hours")
    print(f"   • Average max gap between chats: {df['Max_Gap_Minutes'].mean():.1f} minutes")
    
    print(f"\n📊 Files Generated:")
    print(f"   • detailed_timestamp_analysis.csv")
    print(f"   • login_sessions_comparison.csv")
    print(f"   • detailed_timeline_analysis.png")
    print(f"   • detailed_summary_statistics.png")

if __name__ == "__main__":
    main()
