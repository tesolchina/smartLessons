#!/usr/bin/env python3
"""
Corrected Chatbot Data Analyzer
Properly interprets the Excel structure for goal-setting chatbot study
Created: September 6, 2025
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import json
import re
import os
from datetime import datetime

def load_and_analyze_chatbot_data():
    """Load and analyze the chatbot data with proper column interpretation"""
    
    print("🤖 Chatbot Usage Analysis - Corrected Version")
    print("="*60)
    
    base_dir = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper"
    excel_file = os.path.join(base_dir, "data", "final_report_fulfill.xlsx")
    
    # Manual data extraction using openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        print(f"📊 Processing {ws.max_row-1} students from Excel file...")
        
        # Column mapping based on the headers we saw:
        # A: personnel_id, B: email, C: Dates of chats, D: avg user word count, 
        # E: avg AI word count, F: Numbers of rounds of chats
        
        results = []
        
        for row_num in range(2, ws.max_row + 1):
            row = list(ws[row_num])
            
            # Extract data
            student_id = str(row[0].value) if row[0].value else f"Student_{row_num-1}"
            email = str(row[1].value) if row[1].value else ""
            dates_raw = str(row[2].value) if row[2].value else "[]"
            avg_user_words = row[3].value if row[3].value else 0
            avg_ai_words = row[4].value if row[4].value else 0
            chat_rounds = row[5].value if row[5].value else 0
            
            # Parse dates to count login sessions
            login_count = 1  # Default to 1 login
            try:
                # The dates appear to be in format like "['2024-09-26T08:20:51.529Z']"
                dates_clean = dates_raw.replace("'", '"')  # Fix quotes for JSON
                if dates_clean.startswith('[') and dates_clean.endswith(']'):
                    dates_list = json.loads(dates_clean)
                    # Count unique dates (login days)
                    unique_dates = set()
                    for date_str in dates_list:
                        try:
                            date_part = date_str.split('T')[0]  # Get just the date part
                            unique_dates.add(date_part)
                        except:
                            pass
                    login_count = len(unique_dates) if unique_dates else 1
            except:
                # Fallback: try to count comma-separated dates
                if ',' in dates_raw:
                    login_count = len(dates_raw.split(','))
            
            # Calculate total words (user words * rounds)
            try:
                chat_rounds_num = float(chat_rounds) if chat_rounds else 0
                avg_user_words_num = float(avg_user_words) if avg_user_words else 0
                total_user_words = int(chat_rounds_num * avg_user_words_num)
            except:
                total_user_words = 0
            
            results.append({
                'Student_ID': student_id,
                'Email': email,
                'Login_Count': login_count,
                'Chat_Rounds': int(float(chat_rounds)) if chat_rounds else 0,
                'Avg_User_Words': round(float(avg_user_words), 2) if avg_user_words else 0,
                'Avg_AI_Words': round(float(avg_ai_words), 2) if avg_ai_words else 0,
                'Total_User_Words': total_user_words,
                'Raw_Dates': dates_raw[:50] + "..." if len(dates_raw) > 50 else dates_raw
            })
        
        df = pd.DataFrame(results)
        
        # Filter out rows with no activity
        df = df[df['Chat_Rounds'] > 0]
        
        print(f"✅ Successfully processed {len(df)} active students")
        
        # Display summary statistics
        print(f"\n📊 SUMMARY STATISTICS")
        print(f"{'='*40}")
        print(f"Total active students: {len(df)}")
        print(f"Total chat rounds: {df['Chat_Rounds'].sum()}")
        print(f"Total user words: {df['Total_User_Words'].sum():,}")
        print(f"Average logins per student: {df['Login_Count'].mean():.2f}")
        print(f"Average chat rounds per student: {df['Chat_Rounds'].mean():.2f}")
        print(f"Average user words per student: {df['Total_User_Words'].mean():.2f}")
        print(f"Average user words per chat: {df['Avg_User_Words'].mean():.2f}")
        print(f"Average AI words per response: {df['Avg_AI_Words'].mean():.2f}")
        
        return df
        
    except Exception as e:
        print(f"❌ Error processing data: {e}")
        return None

def create_comprehensive_charts(df, output_dir):
    """Create comprehensive visualization charts"""
    
    print(f"\n📈 Creating comprehensive visualizations...")
    
    # Set up the plotting style
    plt.style.use('default')
    
    # Main analysis chart (2x3 subplot layout)
    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    fig.suptitle('Goal-Setting Chatbot Usage Analysis', fontsize=16, fontweight='bold')
    
    # 1. Chat Rounds Distribution
    axes[0, 0].bar(range(len(df)), df['Chat_Rounds'], color='lightblue', alpha=0.7)
    axes[0, 0].set_title('Chat Rounds per Student')
    axes[0, 0].set_xlabel('Student Index')
    axes[0, 0].set_ylabel('Number of Chat Rounds')
    axes[0, 0].grid(True, alpha=0.3)
    
    # 2. Total User Words Distribution
    axes[0, 1].bar(range(len(df)), df['Total_User_Words'], color='lightgreen', alpha=0.7)
    axes[0, 1].set_title('Total User Words per Student')
    axes[0, 1].set_xlabel('Student Index')
    axes[0, 1].set_ylabel('Total Words')
    axes[0, 1].grid(True, alpha=0.3)
    
    # 3. Average User Words per Chat
    axes[0, 2].bar(range(len(df)), df['Avg_User_Words'], color='salmon', alpha=0.7)
    axes[0, 2].set_title('Average User Words per Chat')
    axes[0, 2].set_xlabel('Student Index')
    axes[0, 2].set_ylabel('Average Words per Message')
    axes[0, 2].grid(True, alpha=0.3)
    
    # 4. Average AI Words per Response
    axes[1, 0].bar(range(len(df)), df['Avg_AI_Words'], color='gold', alpha=0.7)
    axes[1, 0].set_title('Average AI Words per Response')
    axes[1, 0].set_xlabel('Student Index')
    axes[1, 0].set_ylabel('Average Words per AI Response')
    axes[1, 0].grid(True, alpha=0.3)
    
    # 5. Chat Rounds vs Total Words Scatter
    axes[1, 1].scatter(df['Chat_Rounds'], df['Total_User_Words'], 
                       color='purple', alpha=0.6, s=60)
    axes[1, 1].set_title('Chat Rounds vs Total User Words')
    axes[1, 1].set_xlabel('Chat Rounds')
    axes[1, 1].set_ylabel('Total User Words')
    axes[1, 1].grid(True, alpha=0.3)
    
    # Add correlation coefficient
    correlation = df['Chat_Rounds'].corr(df['Total_User_Words'])
    axes[1, 1].text(0.05, 0.95, f'Correlation: {correlation:.3f}', 
                   transform=axes[1, 1].transAxes, fontweight='bold',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 6. User vs AI Words Comparison
    axes[1, 2].scatter(df['Avg_User_Words'], df['Avg_AI_Words'], 
                       color='red', alpha=0.6, s=60)
    axes[1, 2].set_title('User Words vs AI Response Words')
    axes[1, 2].set_xlabel('Average User Words')
    axes[1, 2].set_ylabel('Average AI Words')
    axes[1, 2].grid(True, alpha=0.3)
    
    # Add correlation
    user_ai_correlation = df['Avg_User_Words'].corr(df['Avg_AI_Words'])
    axes[1, 2].text(0.05, 0.95, f'Correlation: {user_ai_correlation:.3f}', 
                   transform=axes[1, 2].transAxes, fontweight='bold',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    plt.tight_layout()
    
    # Save main chart
    main_chart = os.path.join(output_dir, 'comprehensive_chatbot_analysis.png')
    plt.savefig(main_chart, dpi=300, bbox_inches='tight')
    print(f"✅ Main analysis chart saved: {main_chart}")
    plt.close()
    
    # Summary Statistics Chart
    fig, ax = plt.subplots(1, 1, figsize=(12, 8))
    
    # Calculate summary stats
    stats = {
        'Total Students': len(df),
        'Total Chat Rounds': df['Chat_Rounds'].sum(),
        'Total User Words': df['Total_User_Words'].sum(),
        'Avg Rounds/Student': df['Chat_Rounds'].mean(),
        'Avg User Words/Student': df['Total_User_Words'].mean(),
        'Avg User Words/Chat': df['Avg_User_Words'].mean(),
        'Avg AI Words/Response': df['Avg_AI_Words'].mean()
    }
    
    # Create horizontal bar chart
    y_pos = np.arange(len(stats))
    values = list(stats.values())
    labels = list(stats.keys())
    
    bars = ax.barh(y_pos, values, color=['skyblue', 'lightgreen', 'salmon', 'gold', 
                                         'lightcoral', 'plum', 'lightgray'], alpha=0.7)
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels)
    ax.set_xlabel('Count / Average')
    ax.set_title('Chatbot Usage Summary Statistics', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    
    # Add value labels
    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax.text(width + max(values)*0.01, bar.get_y() + bar.get_height()/2, 
                f'{width:.1f}', ha='left', va='center', fontweight='bold')
    
    plt.tight_layout()
    
    summary_chart = os.path.join(output_dir, 'chatbot_summary_statistics.png')
    plt.savefig(summary_chart, dpi=300, bbox_inches='tight')
    print(f"✅ Summary statistics chart saved: {summary_chart}")
    plt.close()
    
    return True

def save_detailed_results(df, output_dir):
    """Save detailed analysis results"""
    
    # Save main results CSV
    main_csv = os.path.join(output_dir, 'chatbot_usage_analysis_detailed.csv')
    df.to_csv(main_csv, index=False)
    print(f"✅ Detailed results saved: {main_csv}")
    
    # Create summary CSV
    summary_data = {
        'Metric': [
            'Total Active Students',
            'Total Chat Rounds',
            'Total User Words',
            'Average Logins per Student',
            'Average Chat Rounds per Student', 
            'Average User Words per Student',
            'Average User Words per Chat',
            'Average AI Words per Response',
            'User-AI Word Correlation',
            'Chat Rounds-Words Correlation'
        ],
        'Value': [
            len(df),
            df['Chat_Rounds'].sum(),
            df['Total_User_Words'].sum(),
            df['Login_Count'].mean(),
            df['Chat_Rounds'].mean(),
            df['Total_User_Words'].mean(),
            df['Avg_User_Words'].mean(),
            df['Avg_AI_Words'].mean(),
            df['Avg_User_Words'].corr(df['Avg_AI_Words']),
            df['Chat_Rounds'].corr(df['Total_User_Words'])
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_csv = os.path.join(output_dir, 'chatbot_summary_metrics.csv')
    summary_df.to_csv(summary_csv, index=False)
    print(f"✅ Summary metrics saved: {summary_csv}")
    
    return True

def display_top_students(df):
    """Display top performing students"""
    
    print(f"\n🏆 TOP STUDENTS ANALYSIS")
    print(f"{'='*50}")
    
    print(f"\n📊 Top 5 by Chat Rounds:")
    top_rounds = df.nlargest(5, 'Chat_Rounds')
    for i, (_, row) in enumerate(top_rounds.iterrows(), 1):
        print(f"   {i}. Student {row['Student_ID'][:10]}...")
        print(f"      Rounds: {row['Chat_Rounds']}, Words: {row['Total_User_Words']}, "
              f"Avg/Chat: {row['Avg_User_Words']:.1f}")
    
    print(f"\n💬 Top 5 by Total Words:")
    top_words = df.nlargest(5, 'Total_User_Words')
    for i, (_, row) in enumerate(top_words.iterrows(), 1):
        print(f"   {i}. Student {row['Student_ID'][:10]}...")
        print(f"      Words: {row['Total_User_Words']}, Rounds: {row['Chat_Rounds']}, "
              f"Avg/Chat: {row['Avg_User_Words']:.1f}")
    
    print(f"\n🗨️ Top 5 by Average Words per Chat:")
    top_avg = df.nlargest(5, 'Avg_User_Words')
    for i, (_, row) in enumerate(top_avg.iterrows(), 1):
        print(f"   {i}. Student {row['Student_ID'][:10]}...")
        print(f"      Avg/Chat: {row['Avg_User_Words']:.1f}, Rounds: {row['Chat_Rounds']}, "
              f"Total: {row['Total_User_Words']}")

def main():
    """Main execution function"""
    
    # Load and analyze data
    df = load_and_analyze_chatbot_data()
    
    if df is None or len(df) == 0:
        print("❌ No data to analyze")
        return
    
    # Display top students
    display_top_students(df)
    
    # Save results
    output_dir = "/Users/simonwang/Documents/Usage/VibeCoding/DailyAssistant/projects/goal-setting-chatbot-paper/data"
    save_detailed_results(df, output_dir)
    
    # Create charts
    create_comprehensive_charts(df, output_dir)
    
    print(f"\n🎉 ANALYSIS COMPLETED SUCCESSFULLY!")
    print(f"📂 All results saved to: {output_dir}/")
    print(f"📊 Files generated:")
    print(f"   • chatbot_usage_analysis_detailed.csv")
    print(f"   • chatbot_summary_metrics.csv") 
    print(f"   • comprehensive_chatbot_analysis.png")
    print(f"   • chatbot_summary_statistics.png")

if __name__ == "__main__":
    main()
